"""XLSX parser — openpyxl. Mỗi sheet → 1 table + 1 H2."""

from __future__ import annotations

from pathlib import Path

from app.pipeline.parsers import ParsedDoc


def parse(path: Path) -> ParsedDoc:
    from openpyxl import load_workbook  # lazy import

    wb = load_workbook(str(path), data_only=True, read_only=True)
    text_parts: list[str] = []
    tables: list[list[list[str]]] = []

    for sheet in wb.worksheets:
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([("" if v is None else str(v)) for v in row])
        rows = [r for r in rows if any(c.strip() for c in r)]
        if not rows:
            continue
        text_parts.append(f"## {sheet.title}")
        tables.append(rows)

    return ParsedDoc(
        text="\n\n".join(text_parts),
        tables=tables,
        metadata={"source_type": "xlsx", "sheets": str(len(wb.worksheets))},
    )
